import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

class ExcelStorage:
    def __init__(self, file_path="attendance.xlsx"):
        self.file_path = file_path
        self.archive_dir = "archive"
        if not os.path.exists(self.archive_dir):
            os.makedirs(self.archive_dir)
        self._check_and_archive_old_month()
        self._initialize_file()

    def _check_and_archive_old_month(self):
        """
        Checks if the current file belongs to a previous month.
        If so, moves it to the archive and a new one will be initialized.
        """
        if not os.path.exists(self.file_path):
            return

        # Get file creation or last modification time
        file_time = datetime.fromtimestamp(os.path.getmtime(self.file_path))
        current_time = datetime.now()

        # If month or year changed, archive it
        if file_time.month != current_time.month or file_time.year != current_time.year:
            archive_name = f"attendance_{file_time.strftime('%Y-%m')}.xlsx"
            archive_path = os.path.join(self.archive_dir, archive_name)
            
            # Close any handles if possible (shutil.move is generally safe)
            try:
                shutil.move(self.file_path, archive_path)
                print(f"Таблица за прошлый месяц архивирована: {archive_name}")
            except Exception as e:
                print(f"Ошибка при архивации: {e}")

    def _initialize_file(self):
        """
        Creates the Excel file if it doesn't exist.
        """
        if not os.path.exists(self.file_path):
            wb = Workbook()
            # Main attendance sheet
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["User ID", "Фамилия", "Action", "Date", "Time"])
            
            # Employees sheet
            wb.create_sheet("Employees")
            emp_ws = wb["Employees"]
            emp_ws.append(["user_id", "Фамилия"])
            
            # Summary sheet (Matrix)
            wb.create_sheet("Summary")
            summary_ws = wb["Summary"]
            summary_ws.append(["Фамилия"])
            
            wb.save(self.file_path)

    def _update_summary(self, full_name, date_str, value="+"):
        """
        Updates the Summary sheet with a value (default '+') for the given employee and date.
        """
        wb = load_workbook(self.file_path)
        if "Summary" not in wb.sheetnames:
            wb.create_sheet("Summary")
            ws = wb["Summary"]
            ws.append(["Фамилия"])
        else:
            ws = wb["Summary"]

        # Find or create the column for the date
        date_col = None
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == date_str:
                date_col = col
                break
        
        if date_col is None:
            date_col = ws.max_column + 1
            cell = ws.cell(row=1, column=date_col)
            cell.value = date_str
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Find or create the row for the employee
        emp_row = None
        # Normalizing name for comparison
        target_name = str(full_name).strip().lower()
        
        for row in range(2, ws.max_row + 1):
            current_cell_value = str(ws.cell(row=row, column=1).value).strip().lower()
            if current_cell_value == target_name:
                emp_row = row
                break
        
        if emp_row is None:
            emp_row = ws.max_row + 1
            ws.cell(row=emp_row, column=1).value = full_name

        # Mark with the value (plus or hours) and center it
        val_cell = ws.cell(row=emp_row, column=date_col)
        val_cell.value = value
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width for the first column (Names)
        ws.column_dimensions['A'].width = 40
        
        # Center the name cell as well
        ws.cell(row=emp_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        # Also center other cells in the matrix for consistency
        for col in range(2, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
        
        wb.save(self.file_path)

    def add_attendance(self, user_id, full_name, action, value="+"):
        """
        Adds attendance record to the Excel file.
        action: 'Пришел' or 'Ушел раньше'
        """
        wb = load_workbook(self.file_path)
        ws = wb["Attendance"]
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        ws.append([user_id, full_name, action, date_str, time_str])
        wb.save(self.file_path)
        
        # Update the summary matrix
        if action in ["Пришел", "Ушел раньше"]:
            print(f"Обновление сводной таблицы для: {full_name} (значение: {value})")
            self._update_summary(full_name, date_str, value)
            
        return True

    def get_employee_name(self, user_id):
        """
        Returns the registered full name for a user_id.
        """
        wb = load_workbook(self.file_path)
        ws = wb["Employees"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_id):
                return row[1]
        return None

    def register_employee(self, user_id, full_name):
        """
        Registers a new employee in the 'Employees' sheet.
        """
        wb = load_workbook(self.file_path)
        ws = wb["Employees"]
        
        # Check if already registered
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_id):
                return False
        
        ws.append([user_id, full_name])
        wb.save(self.file_path)
        
        # Also ensure they have a row in the Summary sheet
        if "Summary" in wb.sheetnames:
            summary_ws = wb["Summary"]
            # Check if name already in Summary
            found = False
            for row in range(2, summary_ws.max_row + 1):
                if summary_ws.cell(row=row, column=1).value == full_name:
                    found = True
                    break
            if not found:
                summary_ws.append([full_name])
                wb.save(self.file_path)
                
        return True
